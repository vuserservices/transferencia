import xlwings as xw
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import win32com.client
import numpy as np
from pandas import ExcelWriter
import time
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential
from office365.sharepoint.files.file import File
import io
from io import BytesIO
from openpyxl.styles import NamedStyle
import win32com.client
import re



site_url = "https://aquanimavdi.sharepoint.com/sites/DDBBCE"
usersp="francisco.alvarez@aquanimaservices.com"
passsp="Aquanima.2023"

def descarga_sp(site_url,usersp,passsp):
    
    ctx = ClientContext(site_url).with_credentials(UserCredential(usersp, passsp))
    file_relative_path = "/sites/DDBBCE/Documentos compartidos/SGTBBDD/Informe pedidos.xls"

    file_relative_path2 = "/sites/DDBBCE/Documentos compartidos/SGTBBDD/Informe solicitudes.zip"
    file_relative_path3 = "/sites/DDBBCE/Documentos compartidos/SGTBBDD/BBDD.xlsx"
    file_relative_path4 = "/sites/DDBBCE/Documentos compartidos/SGTBBDD/informe-SGT.xlsx"

    response = File.open_binary(ctx, file_relative_path)
    bytes_file_obj= io.BytesIO()
    bytes_file_obj.write(response.content)
    bytes_file_obj.seek(0)

    response2= File.open_binary(ctx,file_relative_path2)
    bytes_file_obj1= io.BytesIO()
    bytes_file_obj1.write(response2.content)
    bytes_file_obj1.seek(0)

    response3= File.open_binary(ctx,file_relative_path3)
    bytes_file_obj2= io.BytesIO()
    bytes_file_obj2.write(response3.content)
    bytes_file_obj2.seek(0)
   
    response4= File.open_binary(ctx,file_relative_path4)
    bytes_file_obj4= io.BytesIO()
    bytes_file_obj4.write(response4.content)
    bytes_file_obj4.seek(0)

    with open('./sgt/Seguimiento de pedidos.xls', 'wb') as f:
        f.write(bytes_file_obj.getvalue())
    
    with open('./sgt/archivo.zip', 'wb') as f:
        f.write(bytes_file_obj1.getvalue())
    with open('./sgt/BBDD1.xlsx', 'wb') as f:
        f.write(bytes_file_obj2.getvalue())

    with open('./sgt/INFORME_SGT.xlsx', 'wb') as f:
        f.write(bytes_file_obj4.getvalue())
        pass



start_time = time.time()
print("INICIO")

bbdd="./out_test.xlsx"#plantilla para cargar información
bbdd_sp='./BBDD.xlsx'#traer de Sharepoint
informe_sgt= './sgt/informesgt.xlsx' #traer de sharepoint
contracts_ariba='./sgt/contratos.xlsx'
sourcing_ariba='./sgt/sourcing.xlsx'
df_sourcing_ariba=pd.read_excel(sourcing_ariba)

df_contract_ariba=pd.read_excel(contracts_ariba)
df_contract_ariba['Contract Id'] = df_contract_ariba['Contract Id'].astype(str)


df_sourcing_ariba['Cost Line ID2']=df_sourcing_ariba['Cost Line ID'].replace({'-0000000001': '', '-0000000002': ''}, regex=True)
df_sourcing_ariba['Associated CW2'] = df_sourcing_ariba['Associated CW'].str.split('[', n=1).str[0].str.replace(' ', '')
df_sourcing_ariba= df_sourcing_ariba.drop(columns=['Owner Name'])


print(df_contract_ariba.columns)
df_status_Ariba = pd.merge(df_sourcing_ariba, df_contract_ariba[['Contract Id', 'OLD Buckets','Owner Name','Effective Date - Date','Expiration Date - Date','Supplier - Common Supplier','Supplier Tax ID']], how='left', left_on='Associated CW2', right_on='Contract Id')
#"con este df se busca la columna documentacion de nuevabbddd para las s olicitudes que no tienen contrato escrito dentoer de las bases"
print(df_status_Ariba.columns)
df_status_Ariba = df_status_Ariba[['Cost Line ID2','OLD Buckets','Contract Id','Owner Name','Effective Date - Date','Expiration Date - Date','Supplier - Common Supplier','Supplier Tax ID']]


print(df_status_Ariba.columns)

fila_contrato = df_contract_ariba.loc[df_contract_ariba['Contract Id'] == 'CW267265']
print(fila_contrato)

#region Filtros
control= ["4. CONTRATACIÓN DOTACIÓN DE FONDOS","5D.AUDITORÍA ANUAL DE CUENTAS","5F2.PRESIDENCIA-ALTA DIRECCIÓN","5F3.ESTRATÉGICOS.OPERATIVOS.NEGOCIO","5F3.ESTRATÉGICOS.OPERATIVOS.NEGOCIO",
          "5F3.ESTRATÉGICOS/OPERATIVOS/NEGOCIO","5F3.ESTRATÉGICOS/OPERATIVOS/NEGOCIO ","5F4.FUERA DE PERÍMETRO DE COMPRAS","5F3.ESTRATÉGICOS/OPERATIVOS/NEGOCIO	","5F5.IMPORTE < NEGOCIACIÓN AQUANIMA",
          "5G.INCUMPLIMIENTO","6. BOLSA DE GASTO DISCRECIONAL"
]
control2=['1. COMPRA PUNTUAL','5A.PROVEEDOR INTRAGRUPO','SIN SOLICITUD']

control3= {'01.Elaboración AQN':'4 - Elaboración AQN',	'01. Elaboración':'4 - Elaboración AQN',	'05.Validación Usuario':'7 - Validación Usuario',	'08.Pend. enviar a validación AJ':'8 - Validación de AAJJ',	'09.Validación Asesoría Jurídica':'8 - Validación de AAJJ',	'1. Elaboration':'4 - Elaboración AQN',	'14.Pend. activar/enviar firma Proveedor':'9 - En firma de proveedor',	'6. En firma':'9 - En firma de proveedor',	'15.Firma Proveedor':'9 - En firma de proveedor',	'17.Firma Electrónica':'9 - En firma de proveedor',	'18. Alegaciones de Proveedor':'8 - Alegaciones del proveedor',	'19.Pend. enviar a firma Cliente':'9 - Firma Cliente',	'2. User Validation':'7 - Validación Usuario',	'20.Firma Cliente':'9 - Firma Cliente',	'3. Legal':'8 - Validación de AAJJ',	'5. Supplier Validation':'8 - Revisión proveedor',	'09. En Firma Proveedor':'9 - En firma de proveedor',	'6. Supplier Signature':'9 - En firma de proveedor',	'7. Distribution':'9 - Firmado',	'Alegaciones del proveedor':'8 - Alegaciones del proveedor',	'BORRADA':'1 - Solicitud borrada',	'Cancelado':'2 - Cancelado',	'DEVUELTA':'2 - Solicitud devuelta',	'Dotación':'9 - Dotación',	'Elaboración AQN':'4 - Elaboración AQN',	'Elaboración proveedor':'6 - Elaboración proveedor',	'01. Elaboración proveedor':'6 - Elaboración proveedor',	'Elevado al Foro':'2 - Elevado al Foro',
           	'En negociación':'3 - En negociación',	'Excepción':'1 - Excepción',	'Firma de cliente':'9 - Firma Cliente',	'Firma proveedor':'9 - En firma de proveedor',	'9. En Firma Proveedor':'9 - En firma de proveedor',	'Firmado':'9 - Firmado',	'Gestor del servicio':'7 - Gestor del servicio',	'Intragrupo':'1 - Intragrupo',	'NO NECESARIA GESTIÓN':'1 - No precisa formalización',	'No scope AQN':'1 - No scope AQN',	'ON HOLD':'ON HOLD',	'OT proveedor':'6 - Elaboración proveedor',	'Pdte asignación abogado':'8 - Validación de AAJJ',	'Pdte confirmación cancelación':'1 - Pdte confirmación cancelación',	'Pdte elevar al cliente':'2 - Pte elevar al cliente',	'Pdte elevar al foro':'2 - Pte elevar al foro',	'Pdte homologación':'1 - Pdte homologación',	'Pdte info usuario':'1 - Pdte info usuario',	'Pdte localizar CW':'3 - Pdte localizar CW',	'Pedido no comunicado':'1 - Pedido no comunicado',	'Pendiente de Borrado':'1 - Solicitud borrada',	'Pendiente Marco':'8 - Pediente de Marco',	'Pendiente de MSA':'8 - Pediente de Marco',	'05. Pendiente MSA':'8 - Pediente de Marco',	'Pendiente del Marco':'8 - Pediente de Marco',	'RENOVACIÓN':'1 - Sin información',	'Revisión AQN':'4 - En revisión AQN',	'Pdte info AQN':'4 - En revisión AQN',	'10. Pendiente publicar':'7 - Pendiente publicar',	'8. Pendiente avanzar por propietario':'7 - Pendiente avanzar tarea',
            '08. Pendiente avanzar por propietario':'7 - Pendiente avanzar tarea',	'4. Owner Validation':'7 - Pendiente avanzar tarea',	'04. Validación por proveedor':'8 - Revisión proveedor',	'Revisión proveedor':'8 - Revisión proveedor',	'Sin información':'1 - Sin información',	'02. Revisión por Dpto. Jurídico':'8 - Validación de AAJJ',	'Validación de AAJJ':'8 - Validación de AAJJ',	'Validación de costes':'8 - Validación de costes',	'4. Validación por proveedor':'8 - Revisión proveedor',	'8. Firmado':'9 - Firmado',	'05. Revisión alegaciones por Equipo Datos':'8 - Validación de AAJJ',	'5. Revisión alegaciones por Ciberseguridad':'8 - Validación de AAJJ',	'2. Revisión por Dpto. Jurídico':'8 - Validación de AAJJ',	'0. Initiation':'4 - Elaboración AQN',	'Pendiente de Contrato Marco':'8 - Pediente de Marco',	'Pendiente MSA':'8 - Pediente de Marco',	'Pendiente de firma MSA':'8 - Pediente de Marco',	'06. Validación por Costes':'8 - Validación de costes',	'6. Validación por Costes':'8 - Validación de costes',	'02. Validación usuario':'7 - Validación Usuario',	'VALIDACIÃ“N USUARIO':'7 - Validación Usuario',	'Pendiente de validaciÃ³n por parte de usuario':'7 - Validación Usuario',	'RevisiÃ³n de usuario':'7 - Validación Usuario',	'Validación usuario':'7 - Validación Usuario',	'Pte info Accenture':'1- Pte info Accenture',	'Aclaración':'1 - Aclaración',	'05. Revisión alegaciones por Gestor Servicio':'8 - Validación de AAJJ'
}

hojas_a_mantener_sgt = ["CONTRATOS MARCO SGT", "Histórico Solicitudes", "PROVEEDORES INFORMACION MERCADO","RC","CONTROL", "ARIBA"]
hojas_a_mantener_cliente = ["CONTRATOS MARCO SGT", "RC"]

filtros_dfbbdd2024=  [ '9 - Firmado','8 - Revisión proveedor',  
           '3 - En negociación', '8 - Validación de AAJJ', '1 - Sin información', '4 - En revisión AQN', 
           '7 - Validación Usuario', '6 - Elaboración proveedor', '9 - En firma de proveedor', 
           '2 - Solicitud devuelta', '8 - Validación de costes', '7 - Pendiente avanzar tarea', 
           '4 - Elaboración AQN', '1 - Pedido no comunicado', '1 - Aclaración', '8 - Pediente de Marco', 
            '9 - Firma Cliente', '1 - Pdte homologación', 'ON HOLD']

filtros_dfbbdd2023=  [ '1 - Intragrupo', '1 - Solicitud borrada',
           '2 - Solicitud devuelta',  '2 - Cancelado', '1 - Pedido no comunicado']


def convert_to_float(x):
    try:
        return float(x)
    except (ValueError, TypeError):
        return None

def crear_archivo_SGT(origen, destino, hojas_a_mantener, dfs_a_agregar):
    wb_origen = load_workbook(origen, read_only=True)
    wb_destino = load_workbook(destino)

    for hoja in wb_origen.sheetnames:
        if hoja in hojas_a_mantener:
            ws_origen = wb_origen[hoja]
            ws_destino = wb_destino.create_sheet(title=hoja)

            for row in ws_origen.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        ws_destino[cell.coordinate].value = cell.value

    # Agregar los DataFrames como hojas al libro destino
    for nombre_df, df in dfs_a_agregar.items():
        ws = wb_destino.create_sheet(title=nombre_df)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)


    ws = wb_destino["SOLICITUDES"]
    max_fila = ws.max_row
    max_columna = ws.max_column

    # Definir el rango de celdas para la tabla (excluyendo la fila de encabezados)
    tabla_range = f"A1:{get_column_letter(max_columna)}{max_fila}"

    # Crear un objeto de tabla
    tabla = Table(displayName="DATOS_SOL", ref=tabla_range)

    # Definir el estilo de la tabla
    tabla.style = "TableStyleMedium9"

    # Agregar la tabla a la hoja
    ws.add_table(tabla)
        # Formato de fecha deseado
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')

    columnas_fecha = ['FECHA INICIO SAP', 'FECHA FIN SAP', 'HOY', 'F/ NOTIFICACIÓN\n', 'F/DOCUMENTO SUBSIGUIENTE', 'F/FIN DE ACTIVIDAD',  'F/ APROBACIÓN','FECHA INICIO LAH','FECHA FIN LAH']
    for hoja in wb_destino.sheetnames:
        ws = wb_destino[hoja]
        for columna in columnas_fecha:
            if columna in ws.dimensions:
                for cell in ws[columna]:
                    try:
                        # Intentar convertir el valor de la celda a fecha
                        cell.value = pd.to_datetime(cell.value, errors='coerce')
                        # Aplicar el estilo de fecha corta si la conversión fue exitosa
                        cell.style = date_style
                    except ValueError:

                        pass        

    wb_destino.save('./archivolisto.xlsx')

def crear_archivo_cliente(origen, destino, hojas_a_mantener, dfs_a_agregar):
    wb_origen = load_workbook(origen, read_only=True)
    wb_destino = load_workbook(destino)
    print("comenzando el primer for...")
    for hoja in wb_origen.sheetnames:
        if hoja in hojas_a_mantener:
            ws_origen = wb_origen[hoja]
            ws_destino = wb_destino.create_sheet(title=hoja)

            for row in ws_origen.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        ws_destino[cell.coordinate].value = cell.value
    print("comenznado el segundo for")
    # Agregar los DataFrames como hojas al libro destino
    for nombre_df, df in dfs_a_agregar.items():
        ws = wb_destino.create_sheet(title=nombre_df)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)


    ws = wb_destino["SOLICITUDES"]
    max_fila = ws.max_row
    max_columna = ws.max_column
    print("table range")
    # Definir el rango de celdas para la tabla (excluyendo la fila de encabezados)
    tabla_range = f"A1:{get_column_letter(max_columna)}{max_fila}"

    # Crear un objeto de tabla
    tabla = Table(displayName="DATOS_SOL", ref=tabla_range)

    # Definir el estilo de la tabla
    tabla.style = "TableStyleMedium9"
    print("add table")

    # Agregar la tabla a la hoja
    ws.add_table(tabla)
    wb_destino.save('./archivoclientesexternos.xlsx')

    print("finalizado el archivo... ahora al   xw APP")


def xl_archivocliente():
    
        with xw.App(visible=True) as app:
            try:
                print("iniciando xw")
                wb = xw.Book('./archivoclientesexternos.xlsx')
                # Seleccionar la hoja
                print("selecionando hoja solicitudes xw")
                ws = wb.sheets["SOLICITUDES"]
                print("hoja seleccionada  xw")
                last_row = ws.api.UsedRange.Rows.Count
                print("last row  xw")
                data_range = ws.range("F2:F" + str(last_row))
                data_range.api.AutoFilter(6, ["1. COMPRA PUNTUAL", "4. CONTRATACIÓN DOTACIÓN DE FONDOS"], 7) #7 representa el operador 'igual a'
                print("filtro lsito")
                data_rangeMonto = ws.range("M2:M" + str(last_row))
                data_rangeMonto.api.AutoFilter(13, ">200000")
                print("fin de XW")
                wb.save('./archivoclientesexternos_final.xlsx')    
            
            except Exception as e:
                print("exception:", e)
 
        

def descarga_sharepoint():
    user= "francisco.alvarez@aquanimaservices.com"
    password="Aquanima.2023"

    site_url_contratos = "https://aquanimavdi.sharepoint.com/sites/DDBBCE"
    ctx_contratos = ClientContext(site_url_contratos).with_credentials(UserCredential(user, password))

    path_contratos = "/sites/DDBBCE/Documentos compartidos/contratossgto/Contracts Pipeline_3.0.xlsx"
    response_contratos = File.open_binary(ctx_contratos, path_contratos)

    bytes_file_contratos= io.BytesIO()#nuevo
    bytes_file_contratos.write(response_contratos.content)#nuevo
    bytes_file_contratos.seek(0)#nuevo

    df_contracts = pd.read_excel(bytes_file_contratos,sheet_name="UNIQUE CONTRACTS")





def extract_year(x):
    try:
        return x.split(".")[-1]
    except (AttributeError, ValueError):
        return None
    
def comparar_fechas(row):
    hoy = datetime.now().strftime('%d/%m/%Y')
    if row['FECHA INICIO SAP'] < hoy < row['FECHA FIN SAP']:
        return 1
    else:
        return 0
    
def informe_solicitudes():
    try:   
        print("inicio try")
        i=datetime.now()
        #INFORME FECHAS

        dffechas= pd.read_excel(bbdd_sp,sheet_name="ESTADO MES")
        #dfsolicitudesaqn= pd.read_excel(bbdd_sp,sheet_name="SOLICITUDES AQN")
        dfbbdd= pd.read_excel(bbdd_sp,sheet_name='SOLICITUDES')
        dfrc= pd.read_excel(bbdd_sp,sheet_name="RC")
        dfrc['id servicio'] = dfrc['id servicio'].astype('object')
        idioma_rc= {'Crítico':'Critical','Alto':'High','Medio':'Medium','Bajo':'Low','Sin riesgo':'No Risk','Muy bajo':'Very low'}

        dfrc['RISK'] = dfrc['RISK'].replace(idioma_rc)

        dfrc['Relevance'] = dfrc['Relevance'].replace(idioma_rc)


        wb = load_workbook('BBDDpruebas.xlsx', read_only=True)
        hojas_a_eliminar = ["CRUCE", "SOLICITUDES", "PEDIDOS", "LAH"]
        # Elimina la hoja deseada (reemplaza 'Nombre_Hoja' con el nombre de la hoja que deseas eliminar)
        for nombre_hoja in hojas_a_eliminar:
            if nombre_hoja in wb.sheetnames:
                # Elimina la hoja si existe en el libro de trabajo
                hoja = wb[nombre_hoja]
                wb.remove(hoja)
        wb.close()

        dffechas['Fecha'] = pd.to_datetime(dffechas['Fecha'],format='%d/%m/%Y',errors='coerce')
        dffechas = dffechas.sort_values(by=['SOLICITUD', 'Fecha'], ascending=[True, False])  # Ordenar por "SOLICITUD" y "FECHA" de manera descendente
        dffechas = dffechas.drop_duplicates(subset='SOLICITUD', keep='first')  # Eliminar duplicados, manteniendo solo la primera entrada (la de mayor fecha)
        dffechas= dffechas.astype(str)


  
        #INFORME SOLPED Y OC

        #dfoc=pd.read_html('./sgt/Seguimiento de pedidos.xls')[0] ##cuadno viene como XLS
        dfoc=pd.read_excel('./sgt/Seguimiento de pedidos.xlsx')
        #dfoc = dfoc.iloc[57:]

        #dfoc.columns = dfoc.iloc[0]

        #dfoc = dfoc.iloc[1:]
        print(dfoc.columns)

        dfoc=dfoc[dfoc['Estado Pedido']=="Pedido"]


        #dfsol = pd.read_html('./sgt/Seguimiento de solicitudes.xls')[0] ##cuadno viene como XLS
        dfsol = pd.read_excel('./sgt/Seguimiento de solicitudes.xlsx')
        dfsol = dfsol.iloc[101:]

        dfsol.columns = dfsol.iloc[0]

        dfsol = dfsol.iloc[1:]
        print("POR ACA VA SOLICITUDES...")
        print(dfsol.columns)
        dfsol=dfsol.drop(columns=['Solicitud','Organismo aprobador','Nº Negociación', 'Estado Negociación', 'Nº Contrato'])
        dfsol['Inicio de Actividad'] = dfsol['Inicio de Actividad'].str.replace('.', '/')

        # Reemplazar puntos por barras en 'Fin de Actividad'
        dfsol['Fin de Actividad'] = dfsol['Fin de Actividad'].str.replace('.', '/')
        print(dfsol.columns)


        ########

        i=datetime.now()
        #contratos de sharepoint dashboard LAH
        dfcontract='./sgt/contratos.xlsx' ###CAMBIAR AL ARCHIVO QUE REALMENTE VA A LEER

        dfcontract=pd.read_excel(dfcontract)
        dfcontract= dfcontract.drop(columns=['COMMODITIES', 'Buckets CLAR', 'Stock US', 'Retirar CORP', 'Stock MX', 'Stock BR', 'Associated Team'])

        #HOJA CONTRATOS de archivo sharepoint
        

        df=pd.read_excel(informe_sgt)



        df = df.iloc[0:]


        df.columns = df.iloc[0]
        df = df.iloc[1:]
        df = df.iloc[:, :23]
        df["AÑO"] = df['F/INICIO DE ACTIVIDAD'].apply(extract_year)



        df=df[df['AÑO']=="2024"]
        df=df.rename({'PROPUESTA': 'Propuesta'}, axis=1)

        print(df)
        print(df['F/INICIO DE ACTIVIDAD'])


        ###########
        print("AQUI VA")

        dfbbdd = dfbbdd.iloc[1:]

        dfbbdd.columns = dfbbdd.iloc[0]
        dfbbdd = dfbbdd.iloc[1:]

        dfbbdd_2023=dfbbdd[dfbbdd["AÑO"]!=2024]
        dfbbdd= dfbbdd[dfbbdd["AÑO"]==2024]
        dfbbdd_2023_2 = dfbbdd[dfbbdd['DOCUMENTACION'].isin(filtros_dfbbdd2023)] #del df2024 en bbdd que está OK, así no se modifica
        
        
        dfbbdd = dfbbdd[dfbbdd['DOCUMENTACION'].isin(filtros_dfbbdd2024)]#filtramos todo lo que se modifica


        dfbbdd_2023_final = pd.concat([dfbbdd_2023, dfbbdd_2023_2],ignore_index=True)
        dfbbdd_2023_final = dfbbdd_2023_final.iloc[:, :45]


        




        dfbbdd = dfbbdd.iloc[:, :29]
        dfbbdd = dfbbdd.iloc[:-1]
        old_column_names = ['F/INICIO DE ACTIVIDAD ', 'F/FIN DE ACTIVIDAD ','F. NOTIFICACIÓN\n']
        new_column_names = ['F/INICIO DE ACTIVIDAD', 'F/FIN DE ACTIVIDAD','F/ NOTIFICACIÓN\n']      
        dfbbdd=dfbbdd.rename(columns=dict(zip(old_column_names, new_column_names)))
        columnas_fecha = ['F/ APROBACIÓN','F/INICIO DE ACTIVIDAD', 'F/FIN DE ACTIVIDAD', 'F/DOCUMENTO SUBSIGUIENTE', 'F/ NOTIFICACIÓN\n']



        # Convertir la columna PNETO a tipo float, reemplazando cualquier valor no convertible por 0.0
        dfbbdd['P/NETO'] = dfbbdd['P/NETO'].apply(convert_to_float)
        # Cambiar el formato de las columnas de 'Y/m/d' a 'd/m/Y'
        for columna in columnas_fecha:
            # Convertir a formato de fecha y luego a formato de cadena
            dfbbdd[columna] = pd.to_datetime(dfbbdd[columna], errors='coerce').dt.strftime('%d-%m-%Y')
            # Reemplazar los valores NaN con cadenas vacías

        

        ######

        nuevo_dfbbdd = pd.concat([dfbbdd, df], ignore_index=True)
 


        #nuevo_dfbbdd = nuevo_dfbbdd.drop_duplicates(subset='Propuesta', keep='first')

        print("aqui esta el print despeyus del drop")
        nuevo_dfbbdd.drop(nuevo_dfbbdd.columns[24],axis=1)
        print(nuevo_dfbbdd.columns)
        #nuevo_dfbbdd.insert(0, columna.name, columna)
  
        nuevo_dfbbdd=nuevo_dfbbdd.drop(columns=['PROPUESTA'])




        nuevo_dfbbdd["SOLICITUD000"]=nuevo_dfbbdd['SOLICITUD'].astype(str).str.zfill(10)
        nuevo_dfbbdd=pd.merge(nuevo_dfbbdd, df_status_Ariba[['Cost Line ID2','Contract Id','OLD Buckets','Owner Name']], how='left', left_on='SOLICITUD000', right_on='Cost Line ID2')
        nuevo_dfbbdd=nuevo_dfbbdd.rename({'Owner Name': 'ANALISTA','OLD Buckets':'DOCUMENTACIÓN'}, axis=1)

        





        nuevo_dfbbdd= nuevo_dfbbdd.astype(str)
        nuevo_dfbbdd['SOLICITUD'] = nuevo_dfbbdd['SOLICITUD'].astype(str)
        nuevo_dfbbdd = nuevo_dfbbdd.replace(['nan', 'NaN', 'N/A', np.nan], '')

        
        
        #hacer un merge entre Frechazomesa y ID CONTRAto para los vacíos, luego eliminar FRechazomesa o IDCONTRATO

        nuevo_dfbbdd['DOCUMENTACIÓN'] = np.where(nuevo_dfbbdd['DOCUMENTACIÓN']=='', nuevo_dfbbdd['Contract Id'], nuevo_dfbbdd['DOCUMENTACIÓN'])
        #luego lo mismo con documentación, pero no solo los vacíos.
        nuevo_dfbbdd=nuevo_dfbbdd.drop(columns=['SOLICITUD000', 'Cost Line ID2','Contract Id' ])
        nuevo_dfbbdd['ID CONTRATO']=nuevo_dfbbdd['ID CONTRATO'].astype(str)
        nuevo_dfbbdd=pd.merge(nuevo_dfbbdd,df_contract_ariba[['Contract Id','OLD Buckets','Owner Name','Effective Date - Date', 'Expiration Date - Date']],how='left',left_on='ID CONTRATO' ,right_on='Contract Id')
        
        nuevo_dfbbdd['DOCUMENTACIÓN'] = np.where(nuevo_dfbbdd['DOCUMENTACIÓN']=='' , nuevo_dfbbdd['OLD Buckets'], nuevo_dfbbdd['DOCUMENTACIÓN'])
        nuevo_dfbbdd['ANALISTA'] = np.where(nuevo_dfbbdd['ANALISTA']=='', nuevo_dfbbdd['Owner Name'], nuevo_dfbbdd['ANALISTA'])
        nuevo_dfbbdd['DOCUMENTACIÓN'] = nuevo_dfbbdd['DOCUMENTACIÓN'].replace(control3)
        nuevo_dfbbdd=nuevo_dfbbdd.drop(columns=['Contract Id','OLD Buckets' ])

        


        #####FIltros cuando están vacías las celdas en documentacion:
        

        print('POR ACA')
        

        
        ###############
        #MERGE CON RC
        nuevo_dfbbdd = nuevo_dfbbdd.replace(['nan', 'NaN', 'N/A', np.nan], '')
        nuevo_dfbbdd = pd.merge(nuevo_dfbbdd, dfrc[['WS SARA','RISK','Relevance']], how='left', left_on='CODIGO SERVICIO', right_on='WS SARA')
        nuevo_dfbbdd['Tipologia de Riesgo'] = np.where(nuevo_dfbbdd['Tipologia de Riesgo']=='', nuevo_dfbbdd['RISK'], nuevo_dfbbdd['Tipologia de Riesgo'])
        nuevo_dfbbdd['Tipologia Relavancia'] = np.where(nuevo_dfbbdd['Relevance']=='', nuevo_dfbbdd['Tipologia Relavancia'], nuevo_dfbbdd['Tipologia Relavancia'])


        nuevo_dfbbdd= nuevo_dfbbdd.drop(columns=['RISK','Relevance'])   

        """
        nuevo_dfbbdd['CONTRATOSENCONTRADOS'] = np.nan
        dfcontratos = dfcontratos.astype(str)
        nuevo_dfbbdd.loc[nuevo_dfbbdd['SOLICITUD'].isin(dfcontratos['SOLICITUD']), 'CONTRATOSENCONTRADOS'] = 'OK'
        nuevo_dfbbdd.loc[nuevo_dfbbdd['CONTRATOSENCONTRADOS'].isnull() & nuevo_dfbbdd['TIPO SOLICITUD'].isin(control), 'CONTRATOSENCONTRADOS'] = 'PTE'
        nuevo_dfbbdd.loc[nuevo_dfbbdd['CONTRATOSENCONTRADOS'].isnull() & nuevo_dfbbdd['TIPO SOLICITUD'].isin(control2), 'CONTRATOSENCONTRADOS'] = 'N/A'
        """
       


        #CRUCE CON DFSOL y DFOC
        nuevo_dfbbdd = pd.merge(nuevo_dfbbdd, dfsol[['Nº Solicitud','Estado Solicitud','Inicio de Actividad','Fin de Actividad']], how='left', left_on='SOLICITUD', right_on='Nº Solicitud')
        nuevo_dfbbdd = pd.merge(nuevo_dfbbdd, dfoc[['Nº Solicitud','Proveedor',"N.I.F proveedor"]], how='left', left_on='SOLICITUD', right_on='Nº Solicitud')
        nuevo_dfbbdd = nuevo_dfbbdd.rename(columns={'Inicio de Actividad': 'FECHA INICIO SAP',"Fin de Actividad":"FECHA FIN SAP",'Effective Date - Date':'FECHA INICIO LAH','Expiration Date - Date':'FECHA FIN LAH'})
        nuevo_dfbbdd['FECHA INICIO SAP'] = pd.to_datetime(nuevo_dfbbdd['FECHA INICIO SAP'], format='%d/%m/%Y',errors='coerce')
        print('por aquí')
        
        nuevo_dfbbdd['FECHA FIN SAP'] = pd.to_datetime(nuevo_dfbbdd['FECHA FIN SAP'], format='%d/%m/%Y',errors='coerce')

        nuevo_dfbbdd["HOY"]=datetime.now().date()
        
        nuevo_dfbbdd['En vigor'] = (nuevo_dfbbdd['FECHA INICIO SAP'] < nuevo_dfbbdd["HOY"]) & (nuevo_dfbbdd['HOY'] < nuevo_dfbbdd['FECHA FIN SAP'])
        nuevo_dfbbdd['En vigor'] = nuevo_dfbbdd['En vigor'].astype(int)
        nuevo_dfbbdd['Solicitud Borrada'] = np.where(nuevo_dfbbdd['Estado Solicitud'] == 'Borrado', 'SI', 'NO')
        print("solicitudes borradas")
        ####MERGE CONE STADO MES:
  
        nuevo_dfbbdd = pd.merge(nuevo_dfbbdd, dffechas[['SOLICITUD','DOCUMENTACION']], how='left', left_on='SOLICITUD', right_on='SOLICITUD')
        print(' LUEGO DEL MERGE CON DFFECHAS')
        nuevo_dfbbdd = nuevo_dfbbdd.replace(['nan', 'NaN', 'N/A', np.nan], '')
        nuevo_dfbbdd['DOCUMENTACIÓN'] = np.where(nuevo_dfbbdd['DOCUMENTACIÓN']=='', nuevo_dfbbdd['DOCUMENTACION'], nuevo_dfbbdd['DOCUMENTACIÓN'])
        nuevo_dfbbdd = nuevo_dfbbdd.replace(['nan', 'NaN', 'N/A', np.nan], '')
        nuevo_dfbbdd= nuevo_dfbbdd.drop(columns=['DOCUMENTACION'])
        
        print('LUEGO DEL DROP DOCUMENTAIOCN')



        nuevo_dfbbdd['ID CONTRATO'] = nuevo_dfbbdd.apply(lambda row: re.search(r'(?:C|c)(?:W|w)\s?\d{5,6}', row['CONTRATO']).group() if (pd.isna(row['ID CONTRATO']) or row['ID CONTRATO']=='') and re.search(r'(?:C|c)(?:W|w)\s?\d{5,6}', row['CONTRATO'])!=None else row['ID CONTRATO'] , axis=1)
        nuevo_dfbbdd['ID CONTRATO'] = nuevo_dfbbdd.apply(lambda row: re.search(r'(?:C|c)(?:W|w)\s?\d{5,6}', row['NOTA INTERNA']).group() if (pd.isna(row['ID CONTRATO']) or row['ID CONTRATO']=='') and re.search(r'(?:C|c)(?:W|w)\s?\d{5,6}', row['NOTA INTERNA'])!=None else row['ID CONTRATO'] , axis=1)
        
        print("aquio está el contrato en el nuevodffbb para que lo encuentr........")
        fila_contrato = nuevo_dfbbdd.loc[nuevo_dfbbdd['ID CONTRATO'] == 'CW267265']
        print(fila_contrato)
        print("AQUI EL PRINT DE COLUMNS DE NUEVODFFBBSDD")
        print(nuevo_dfbbdd.columns)


        condicion = (nuevo_dfbbdd['DOCUMENTACIÓN'] == '') & (nuevo_dfbbdd['ID CONTRATO'] != '') 


        filas_a_actualizar = nuevo_dfbbdd.loc[condicion]

        for indice, fila in filas_a_actualizar.iterrows():
            id_contrato = fila['ID CONTRATO']
            fila_status_ariba = df_status_Ariba[df_status_Ariba['Contract Id'] == id_contrato]

            if not fila_status_ariba.empty:
                nuevo_valor_doc = fila_status_ariba['OLD Buckets'].values[0]

                # Actualizar la columna 'DOCUMENTACIÓN'
                nuevo_dfbbdd.at[indice, 'DOCUMENTACIÓN'] = nuevo_valor_doc

                # Obtener los nuevos valores para las otras columnas

                nuevo_inicio_lah = fila_status_ariba['Effective Date - Date'].values[0]
                nuevo_fin_lah = fila_status_ariba['Expiration Date - Date'].values[0]
                nuevo_owner = fila_status_ariba['Owner Name'].values[0]
                nuevo_cif = fila_status_ariba['Supplier Tax ID'].values[0]
                nuevo_nombre = fila_status_ariba['Supplier - Common Supplier'].values[0]

                # Actualizar las otras columnas

                nuevo_dfbbdd.at[indice, 'FECHA INICIO LAH'] = nuevo_inicio_lah
                nuevo_dfbbdd.at[indice, 'FECHA FIN LAH'] = nuevo_fin_lah
                nuevo_dfbbdd.at[indice, 'ANALISTA'] = nuevo_owner
                nuevo_dfbbdd.at[indice, 'N.I.F proveedor'] = nuevo_cif
                nuevo_dfbbdd.at[indice, 'Proveedor'] = nuevo_nombre
                
            
            else:
                # Si no hay coincidencia, dejar los valores originales en las otras columnas
                pass  


        for indice, fila in filas_a_actualizar.iterrows():
            id_contrato = fila['ID CONTRATO']
            fila_contract_ariba = df_contract_ariba[df_contract_ariba['Contract Id'] == id_contrato]

            if not fila_contract_ariba.empty:
                nuevo_valor_doc = fila_contract_ariba['OLD Buckets'].values[0]

                # Actualizar la columna 'DOCUMENTACIÓN'
                nuevo_dfbbdd.at[indice, 'DOCUMENTACIÓN'] = nuevo_valor_doc

                # Obtener los nuevos valores para las otras columnas

                nuevo_inicio_lah = fila_contract_ariba['Effective Date - Date'].values[0]
                nuevo_fin_lah = fila_contract_ariba['Expiration Date - Date'].values[0]
                nuevo_owner = fila_contract_ariba['Owner Name'].values[0]
                nuevo_cif = fila_contract_ariba['Supplier Tax ID'].values[0]
                nuevo_nombre = fila_contract_ariba['Supplier - Common Supplier'].values[0]

                # Actualizar las otras columnas

                nuevo_dfbbdd.at[indice, 'FECHA INICIO LAH'] = nuevo_inicio_lah
                nuevo_dfbbdd.at[indice, 'FECHA FIN LAH'] = nuevo_fin_lah
                nuevo_dfbbdd.at[indice, 'ANALISTA'] = nuevo_owner
                nuevo_dfbbdd.at[indice, 'N.I.F proveedor'] = nuevo_cif
                nuevo_dfbbdd.at[indice, 'Proveedor'] = nuevo_nombre
            
            else:
                # Si no hay coincidencia, dejar los valores originales en las otras columnas
                pass  


        for indice, fila in filas_a_actualizar.iterrows():
            
            id_contrato = fila['ID CONTRATO']
            nuevo_valor = df_contract_ariba.loc[df_contract_ariba['Contract Id'] == id_contrato, 'OLD Buckets'].values
            nuevo_dfbbdd.at[indice, 'DOCUMENTACIÓN'] = nuevo_valor[0] if len(nuevo_valor) > 0 else ''

        #antiguos
        nuevo_dfbbdd['DOCUMENTACIÓN'] = nuevo_dfbbdd.apply(lambda row: '1 - Intragrupo' if row['TIPO SOLICITUD'] == '5A.PROVEEDOR INTRAGRUPO' else row['DOCUMENTACIÓN'], axis=1)
        nuevo_dfbbdd['DOCUMENTACIÓN'] = nuevo_dfbbdd.apply(lambda row: '4 - Elaboración AQN' if (row['DOCUMENTACIÓN'] == '3 - En negociación' or row['DOCUMENTACIÓN'] == '1 - No scope AQN') and pd.notna(row['PEDIDO']) and (pd.isna(row['ID CONTRATO'] or row['ID CONTRATO']=='')) else row['DOCUMENTACIÓN'], axis=1)
        nuevo_dfbbdd['DOCUMENTACIÓN'] = nuevo_dfbbdd.apply(lambda row: '4 - Elaboración AQN' if  pd.notna(row['PEDIDO']) and (pd.isna(row['ID CONTRATO']) or row['ID CONTRATO']=='') and row['TIPO SOLICITUD'] in ['1. COMPRA PUNTUAL,''2. CONTRATACIÓN NUEVO PROD.SERVICIO'] else row['DOCUMENTACIÓN'], axis=1)

        print("por aquioiii")
        #nuevos
        nuevo_dfbbdd['DOCUMENTACIÓN'] = nuevo_dfbbdd.apply(lambda row: '1 - Pdte homologación' if  row['MOTIVO \nRETENCIÓN'] == 'HOMOLOG. GASTO' else row['DOCUMENTACIÓN'], axis=1)
        nuevo_dfbbdd['DOCUMENTACIÓN'] = nuevo_dfbbdd.apply(lambda row: '1 - Solicitud borrada' if row['Estado Solicitud'] == 'Borrado' else row['DOCUMENTACIÓN'], axis=1)
        nuevo_dfbbdd['DOCUMENTACIÓN'] = nuevo_dfbbdd.apply(lambda row: '1 - Sin información' if (pd.isna(row['DOCUMENTACIÓN']) or row['DOCUMENTACIÓN']=='') and row['MOTIVO \nRETENCIÓN']== 'COMUNICADO' and row['TIPO SOLICITUD'] in ['5F3.ESTRATÉGICOS.OPERATIVOS.NEGOCIO', '5F5.IMPORTE < NEGOCIACIÓN AQUANIMA', '4. CONTRATACIÓN DOTACIÓN DE FONDOS', '5F3.ESTRATÉGICOS/OPERATIVOS/NEGOCIO', '5G.INCUMPLIMIENTO', '6. BOLSA DE GASTO DISCRECIONAL']  else row['DOCUMENTACIÓN'], axis=1)
        nuevo_dfbbdd['DOCUMENTACIÓN'] = nuevo_dfbbdd.apply(lambda row: '3 - En negociación' if (pd.isna(row['DOCUMENTACIÓN']) or row['DOCUMENTACIÓN']=='') and row['MOTIVO \nRETENCIÓN'] == 'EN NEGOCIACION' else row['DOCUMENTACIÓN'], axis=1)
        #nuevo_dfbbdd['DOCUMENTACIÓN'] = nuevo_dfbbdd.apply(lambda row: '1 - Sin información' if (pd.isna(row['DOCUMENTACIÓN']) or row['DOCUMENTACIÓN']=='') and row['TIPO SOLICITUD'] == '1. COMPRA PUNTUAL' else row['DOCUMENTACIÓN'], axis=1)
        nuevo_dfbbdd['DOCUMENTACIÓN'] = nuevo_dfbbdd.apply(lambda row: '1 - Pedido no comunicado' if (pd.isna(row['DOCUMENTACIÓN']) or row['DOCUMENTACIÓN']=='')  else row['DOCUMENTACIÓN'], axis=1)

        nuevo_dfbbdd['DOCUMENTACIÓN'] = nuevo_dfbbdd['DOCUMENTACIÓN'].replace(control3)



        #buscar contrato s en columna NOTA INTERNA o CONTrato
        nuevo_dfbbdd["COUNT"]=1
        nuevo_dfbbdd=nuevo_dfbbdd.rename({'PROPUESTA': 'Propuesta'}, axis=1)#,'PROVEEDOR3':'PROVEEDOR'

        nuevo_dfbbdd= nuevo_dfbbdd.drop(columns=['Nº Solicitud_y','Nº Solicitud_x','HOY','Owner Name', 'WS SARA'])

        nuevo_dfbbdd = nuevo_dfbbdd.drop_duplicates(subset="SOLICITUD", keep="first")

        #se une con la base de bbdd
        old_column_names_FINAL = ['N.I.F proveedor','Proveedor','Estado Solicitud','DOCUMENTACIÓN','F/ NOTIFICACIÓN\n','F/INICIO DE ACTIVIDAD','F/FIN DE ACTIVIDAD']
        new_column_names_FINAL = ['CIF Proveedor', 'Nombre Proveedor','Estado SAP','DOCUMENTACION','F. NOTIFICACIÓN\n','F/INICIO DE ACTIVIDAD ','F/FIN DE ACTIVIDAD ']      
        nuevo_dfbbdd=nuevo_dfbbdd.rename(columns=dict(zip(old_column_names_FINAL, new_column_names_FINAL)))
        """
        nuevo_dfbbdd['FECHA INICIO LAH'] = nuevo_dfbbdd['FECHA INICIO LAH'].replace('', pd.NA)
        nuevo_dfbbdd['FECHA FIN LAH'] = nuevo_dfbbdd['FECHA FIN LAH'].replace('', pd.NA)

        # Convertir la columna 'FECHA INICIO LAH' a formato datetime
        nuevo_dfbbdd['FECHA INICIO LAH'] = pd.to_datetime(nuevo_dfbbdd['FECHA INICIO LAH'], errors='coerce', unit='D', origin='1899-12-30')

        # Convertir la columna 'FECHA FIN LAH' a formato datetime
        nuevo_dfbbdd['FECHA FIN LAH'] = pd.to_datetime(nuevo_dfbbdd['FECHA FIN LAH'], errors='coerce', unit='D', origin='1899-12-30')
        """
        print('5')
        print(nuevo_dfbbdd.columns)
        print(dfbbdd_2023_final.columns)

        nuevo_dfbbdd_final=pd.concat([dfbbdd_2023_final,nuevo_dfbbdd],ignore_index=True)
        nuevo_dfbbdd_final= nuevo_dfbbdd_final.drop(columns=['PROPUESTA'])

        print('AQUI PASANDO A EXCEL CON INDEX FALSE')
        nuevo_dfbbdd_final.to_excel('archivolisto_excel.xlsx',index=False)

        print("creado el archivo Excel")
        print('aqui hay un time sleep')

        dfs_a_agregar_sgt = {"SOLICITUDES": nuevo_dfbbdd_final, "CRUCE": dfsol, "PEDIDOS": dfoc, "ESTADO MES":dffechas}#"LAH": dfcontratos,
        archivo_origen_sgt = 'BBDD.xlsx'
        archivo_destino_sgt = 'plantilla.xlsx'
        crear_archivo_SGT(archivo_origen_sgt, archivo_destino_sgt, hojas_a_mantener_sgt, dfs_a_agregar_sgt)

        dfs_a_agregar_cliente = {"CRUCE": dfsol, "SOLICITUDES": nuevo_dfbbdd_final}#"LAH": dfcontratos,
        archivo_origen_cliente = 'BBDDpruebas.xlsx'
        archivo_destino_cliente = 'plantilla.xlsx'
        #crear_archivo_cliente(archivo_origen_cliente,archivo_destino_cliente,hojas_a_mantener_cliente,dfs_a_agregar_cliente)

        print(f'Done! Terminado en {round(time.time()-start_time,2)} segundos.')
    except Exception as e:
        print("El Error fue en Solicitudes: ", e)




######CUANDO ESTE TODO OK, GUARDAR ESTA BASE EN SP Y HACER UNA COPIA ELIMINANDO TODAS LAS HOJAS, EXCEPTO
### DASHBOARD SGT
### CONTRATOS MARCCO SGT
### SOLICITUDES
### SOURCING
### RC




#descarga_sp(site_url,usersp,passsp)
informe_solicitudes()
#xl_archivocliente()
print(f'Tiempo total:', time.time()-start_time)
"""
os.remove("borrar ArchivoOC")
os.remove("borrar ArchivoSOL")
os.remove("Borrar Contratos")
os.remove("Borrar ArchivoSGT")

"""





#una vez listo el informe, pegar hoja a hoja lo que le sirve al cliente y con openpyxl crear las tablas para que lo tome la tabla dinamica de la hoja "dsb"