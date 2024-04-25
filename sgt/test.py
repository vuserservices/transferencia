

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import time
# Cargar el archivo de origen

start_time = time.time()
# Crear un nuevo libro de trabajo

hojas_a_mantener = ["CONTRATOS MARCO SGT", "SOLICITUDES AQN", "Histórico Solicitudes", "PROVEEDORES INFORMACION MERCADO","RC","CONTROL", "ARIBA"]
wb_destino = load_workbook('plantilla.xlsx')  
def copiar_hojas(origen, destino, hojas_a_mantener):
    wb_origen = load_workbook(origen, read_only=True)
    

    for hoja in wb_origen.sheetnames:
        if hoja in hojas_a_mantener:
            ws_origen = wb_origen[hoja]
            ws_destino = wb_destino.create_sheet(title=hoja)

            for row in ws_origen.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        ws_destino[cell.coordinate].value = cell.value

    



# Guardar el archivo de destino
archivo_origen = 'BBDDpruebas.xlsx'
archivo_destino = 'BBDDpruebas_mantenidas.xlsx'

copiar_hojas(archivo_origen, archivo_destino, hojas_a_mantener)

wb_destino.save(wb_destino)
print(f'Done! Completed in {round(time.time()-start_time,2)} seconds.')